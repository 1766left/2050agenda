import json
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def convert_json_to_excel(json_file, excel_file):
    # 读取JSON文件
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 创建一个空的DataFrame
    df = pd.DataFrame()
    
    # 创建一个Excel写入器
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    
    # 写入空的DataFrame以创建Excel文件
    df.to_excel(writer, index=False)
    
    # 获取工作簿和工作表
    workbook = writer.book
    worksheet = workbook.active
    worksheet.title = "活动安排"
    
    # 设置列标题
    headers = ["地点", "时间段", "序号", "时间", "大标题", "环节标题", "姓名", "简介"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    worksheet.column_dimensions['A'].width = 15  # 地点
    worksheet.column_dimensions['B'].width = 15  # 时间段
    worksheet.column_dimensions['C'].width = 8   # 序号
    worksheet.column_dimensions['D'].width = 15  # 时间
    worksheet.column_dimensions['E'].width = 30  # 大标题
    worksheet.column_dimensions['F'].width = 25  # 环节标题
    worksheet.column_dimensions['G'].width = 15  # 姓名
    worksheet.column_dimensions['H'].width = 30  # 简介
    
    # 设置边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 当前行号
    row_num = 2
    
    # 遍历JSON数据并填充Excel
    for location_data in data:
        location = location_data["地点"]
        time_period = location_data["时间段"]
        
        # 记录时间段的起始行
        time_period_start_row = row_num
        
        for activity in location_data["活动"]:
            time = activity.get("时间", "")
            number = activity.get("序号", "")
            title = activity.get("标题", "")
            conveners = activity.get("召集人", [])
            sessions = activity.get("环节", [])
            
            # 记录活动的起始行
            activity_start_row = row_num
            
            # 填充大标题
            title_cell = worksheet.cell(row=row_num, column=5)
            title_cell.value = title
            title_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 填充序号和时间（交换位置）
            number_cell = worksheet.cell(row=row_num, column=3)
            number_cell.value = number
            number_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            time_cell = worksheet.cell(row=row_num, column=4)
            time_cell.value = time
            time_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 处理召集人信息
            if conveners:
                for i, convener in enumerate(conveners):
                    # 修改这里的条件，避免在第一个召集人时多加一行
                    if i > 0:
                        row_num += 1
                    
                    convener_name = convener.get("姓名", "")
                    convener_intro = convener.get("简介", "")
                    
                    # 填充环节标题为"召集人"
                    session_title_cell = worksheet.cell(row=row_num, column=6)
                    session_title_cell.value = "召集人"
                    session_title_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                    
                    # 填充姓名和简介
                    name_cell = worksheet.cell(row=row_num, column=7)
                    name_cell.value = convener_name
                    name_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    intro_cell = worksheet.cell(row=row_num, column=8)
                    intro_cell.value = convener_intro
                    intro_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 处理环节信息
            if sessions:
                for session in sessions:
                    session_title = session.get("标题", "")
                    personnel = session.get("人员", [])
                    
                    # 如果没有人员，至少添加一行
                    if not personnel:
                        personnel = [{"姓名": "", "简介": ""}]
                    
                    # 记录环节标题的起始行
                    session_title_start_row = row_num + 1
                    
                    for i, person in enumerate(personnel):
                        row_num += 1
                        
                        person_name = person.get("姓名", "")
                        person_intro = person.get("简介", "")
                        
                        # 填充环节标题
                        session_title_cell = worksheet.cell(row=row_num, column=6)
                        if i == 0:
                            session_title_cell.value = session_title
                            session_title_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        # 填充人员信息
                        person_name_cell = worksheet.cell(row=row_num, column=7)
                        person_name_cell.value = person_name
                        person_name_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        person_intro_cell = worksheet.cell(row=row_num, column=8)
                        person_intro_cell.value = person_intro
                        person_intro_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # 如果环节有多个人员，合并环节标题单元格
                    if len(personnel) > 1:
                        worksheet.merge_cells(start_row=session_title_start_row, start_column=6, 
                                             end_row=session_title_start_row + len(personnel) - 1, end_column=6)
            
            # 计算实际使用的行数
            actual_rows = row_num - activity_start_row + 1
            
            # 合并单元格
            if actual_rows > 1:
                # 合并序号、时间和大标题单元格
                worksheet.merge_cells(start_row=activity_start_row, start_column=3, 
                                     end_row=row_num, end_column=3)
                worksheet.merge_cells(start_row=activity_start_row, start_column=4, 
                                     end_row=row_num, end_column=4)
                worksheet.merge_cells(start_row=activity_start_row, start_column=5, 
                                     end_row=row_num, end_column=5)
            
            # 移动到下一个活动的起始行
            row_num += 1
        
        # 合并地点和时间段单元格
        worksheet.merge_cells(start_row=time_period_start_row, start_column=1, 
                             end_row=row_num - 1, end_column=1)
        location_cell = worksheet.cell(row=time_period_start_row, column=1)
        location_cell.value = location
        location_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        worksheet.merge_cells(start_row=time_period_start_row, start_column=2, 
                             end_row=row_num - 1, end_column=2)
        time_period_cell = worksheet.cell(row=time_period_start_row, column=2)
        time_period_cell.value = time_period
        time_period_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 应用边框样式到所有单元格
    for row in worksheet.iter_rows(min_row=1, max_row=row_num-1, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border
    
    # 保存Excel文件
    writer.close()
    
    print(f"转换完成，已保存到 {excel_file}")

# 使用示例
if __name__ == "__main__":
    convert_json_to_excel("merged_activities.json", "活动安排.xlsx")